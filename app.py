from flask import Flask, jsonify, request, render_template
import os
import json
from openpyxl import load_workbook, Workbook
from datetime import datetime
from collections import Counter, defaultdict

# Optional PostgreSQL support
try:
    import psycopg2
    PSYCOPG2_AVAILABLE = True
except ImportError:
    PSYCOPG2_AVAILABLE = False

app = Flask(__name__, static_folder='static', template_folder='templates')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
machines_file  = os.path.join(BASE_DIR, "machines.txt")
faults_file    = os.path.join(BASE_DIR, "faults.txt")
remedies_file  = os.path.join(BASE_DIR, "remedies.txt")
employees_file = os.path.join(BASE_DIR, "employees.txt")
parts_file     = os.path.join(BASE_DIR, "parts.txt")
file_name      = os.path.join(BASE_DIR, "machine_log.xlsx")

ADMIN_PASSWORD = "Am1122"

# ─── Helper Functions ────────────────────────────────────────────────────────

def ensure_file(fp):
    if not os.path.exists(fp):
        open(fp, 'w').close()

def load_list(fp):
    ensure_file(fp)
    with open(fp, 'r', encoding='utf-8') as f:
        return [l.strip() for l in f if l.strip()]

def save_list(fp, items):
    with open(fp, 'w', encoding='utf-8') as f:
        f.writelines(f"{i}\n" for i in items)

def initialize_excel():
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Machine Log"
        ws.append(["Machine Name", "Fault", "Remedy", "Shift", "Employee Name",
                   "Fault Start Time", "Fault End Time", "Downtime", "Entry Timestamp",
                   "Machine Status", "Part Change", "Remarks", "Fault Location", "Edit Count"])
        wb.save(file_name)
    else:
        wb = load_workbook(file_name)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        modified = False
        if "Fault Location" not in headers:
            headers.append("Fault Location")
            ws.cell(row=1, column=len(headers), value="Fault Location")
            modified = True
        if "Edit Count" not in headers:
            headers.append("Edit Count")
            ws.cell(row=1, column=len(headers), value="Edit Count")
            modified = True
        if modified:
            wb.save(file_name)

initialize_excel()

def parse_dt(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except ValueError:
            continue
    return None

def verify_admin(req):
    auth = req.headers.get("Authorization", "")
    if auth == f"Bearer {ADMIN_PASSWORD}":
        return True
    data = req.get_json(silent=True) or {}
    return data.get("password") == ADMIN_PASSWORD

# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route('/')
def home():
    return render_template('index.html')

# GET /api/config
@app.route('/api/config', methods=['GET'])
def get_config():
    return jsonify({
        "machines":  load_list(machines_file),
        "faults":    load_list(faults_file),
        "remedies":  load_list(remedies_file),
        "employees": load_list(employees_file),
        "parts":     load_list(parts_file),
    })

# POST /api/config  — add an item
@app.route('/api/config', methods=['POST'])
def add_config_item():
    if not verify_admin(request):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.json or {}
    list_type = data.get("type", "")
    value     = data.get("value", "").strip()
    FILE_MAP = {"machines": machines_file, "faults": faults_file,
                "remedies": remedies_file, "employees": employees_file, "parts": parts_file}
    if list_type not in FILE_MAP or not value:
        return jsonify({"error": "Invalid params"}), 400
    items = load_list(FILE_MAP[list_type])
    if value not in items:
        items.append(value)
        save_list(FILE_MAP[list_type], items)
    return jsonify({"success": True, "items": items})

# DELETE /api/config — remove an item
@app.route('/api/config', methods=['DELETE'])
def delete_config_item():
    if not verify_admin(request):
        return jsonify({"error": "Unauthorized"}), 401
    data = request.json or {}
    list_type = data.get("type", "")
    value     = data.get("value", "").strip()
    FILE_MAP = {"machines": machines_file, "faults": faults_file,
                "remedies": remedies_file, "employees": employees_file, "parts": parts_file}
    if list_type not in FILE_MAP or not value:
        return jsonify({"error": "Invalid params"}), 400
    items = load_list(FILE_MAP[list_type])
    if value not in items:
        return jsonify({"error": "Not found"}), 404
    items.remove(value)
    save_list(FILE_MAP[list_type], items)
    return jsonify({"success": True, "items": items})

# POST /api/admin/login
@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.json or {}
    if data.get("password") == ADMIN_PASSWORD:
        return jsonify({"success": True, "token": ADMIN_PASSWORD})
    return jsonify({"success": False, "error": "Invalid password"}), 401

# POST /api/logs
@app.route('/api/logs', methods=['POST'])
def save_log():
    try:
        data       = request.json or {}
        machine    = data.get("machine",    "").strip()
        fault      = data.get("fault",      "").strip()
        remedy     = data.get("remedy",     "").strip()
        shift      = data.get("shift",      "").strip()
        employee   = data.get("employee",   "").strip()
        start_str  = data.get("startTime",  "").strip()
        end_str    = data.get("endTime",    "").strip()
        status     = data.get("status",     "Running").strip()
        part_change= data.get("partChange", "").strip()
        remarks    = data.get("remarks",    "").strip()

        if not all([machine, fault, remedy, shift, employee, start_str, end_str]):
            return jsonify({"error": "Please fill all required fields."}), 400

        fmt = "%Y-%m-%d %H:%M"
        try:
            start_dt = datetime.strptime(start_str, fmt)
            end_dt   = datetime.strptime(end_str,   fmt)
        except Exception:
            return jsonify({"error": "Invalid date/time format. Use YYYY-MM-DD HH:MM."}), 400

        if end_dt < start_dt:
            return jsonify({"error": "End time cannot be before start time."}), 400

        downtime  = round((end_dt - start_dt).total_seconds() / 60, 2)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([machine, fault, remedy, shift, employee,
                   start_dt.strftime(fmt), end_dt.strftime(fmt),
                   downtime, timestamp, status, part_change, remarks, None, 0])
        wb.save(file_name)

        # Optional Postgres sync
        db_synced = False
        db_error  = None
        if PSYCOPG2_AVAILABLE:
            try:
                conn = psycopg2.connect(
                    host="172.25.0.155", port=5432, database="Shed1",
                    user="postgres", password="12345678", connect_timeout=3
                )
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO shed1_log (
                        machinename, fault, remedy, shift, employeename,
                        faultstarttime, faultendtime, downtime, entrytimestamp,
                        machinestatus, partchange, remarks
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (machine, fault, remedy, shift, employee,
                      start_dt.strftime(fmt), end_dt.strftime(fmt),
                      downtime, timestamp, status, part_change, remarks))
                conn.commit()
                conn.close()
                db_synced = True
            except Exception as e:
                db_error = str(e)

        return jsonify({
            "success": True,
            "db_synced": db_synced,
            "db_error": db_error,
            "entry": {
                "machine": machine, "fault": fault, "remedy": remedy,
                "shift": shift, "employee": employee, "downtime": downtime,
                "status": status, "partChange": part_change,
                "remarks": remarks, "timestamp": timestamp,
                "startTime": start_str, "endTime": end_str
            }
        })
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500

requests_file = os.path.join(BASE_DIR, "requests.json")

def load_requests():
    if not os.path.exists(requests_file):
        return []
    try:
        with open(requests_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []

def save_requests(requests):
    try:
        with open(requests_file, 'w', encoding='utf-8') as f:
            json.dump(requests, f, indent=4)
    except Exception as e:
        print("Error saving requests:", e)

def get_log_by_timestamp(timestamp):
    if not os.path.exists(file_name):
        return None
    try:
        wb = load_workbook(file_name, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        headers = rows[0] if rows else None
        if not headers:
            return None
        hi = {h: i for i, h in enumerate(headers) if h}
        ts_idx = hi.get("Entry Timestamp")
        if ts_idx is None:
            return None
        for row in rows[1:]:
            if not row or len(row) <= ts_idx:
                continue
            if str(row[ts_idx]).strip() == timestamp.strip():
                def gv(row, col, idx):
                    pos = hi.get(col, idx)
                    return row[pos] if pos < len(row) else None
                return {
                    "machine":    gv(row, "Machine Name",    0),
                    "fault":      gv(row, "Fault",           1),
                    "remedy":     gv(row, "Remedy",          2),
                    "shift":      gv(row, "Shift",           3),
                    "employee":   gv(row, "Employee Name",   4),
                    "startTime":  gv(row, "Fault Start Time",5),
                    "endTime":    gv(row, "Fault End Time",  6),
                    "downtime":   gv(row, "Downtime",        7),
                    "timestamp":  timestamp,
                    "status":     gv(row, "Machine Status",  9),
                    "partChange": gv(row, "Part Change",    10),
                    "remarks":    gv(row, "Remarks",        11),
                    "editCount":  int(gv(row, "Edit Count", 13) or 0)
                }
        return None
    except Exception as e:
        print("Error finding log:", e)
        return None

def perform_delete_log(timestamp):
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        try:
            ts_idx = headers.index("Entry Timestamp")
        except ValueError:
            return False, "Entry Timestamp column not found", False, None
            
        row_found = None
        for r_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r_idx, column=ts_idx + 1).value
            if cell_val and str(cell_val).strip() == timestamp:
                row_found = r_idx
                break
                
        if not row_found:
            return False, "Log entry not found", False, None
            
        ws.delete_rows(row_found)
        wb.save(file_name)
        
        # Optional Postgres sync
        db_synced = False
        db_error = None
        if PSYCOPG2_AVAILABLE:
            try:
                conn = psycopg2.connect(
                    host="172.25.0.155", port=5432, database="Shed1",
                    user="postgres", password="12345678", connect_timeout=3
                )
                cur = conn.cursor()
                cur.execute("DELETE FROM shed1_log WHERE entrytimestamp = %s", (timestamp,))
                conn.commit()
                conn.close()
                db_synced = True
            except Exception as e:
                db_error = str(e)
                
        return True, None, db_synced, db_error
    except Exception as e:
        return False, str(e), False, None

def perform_edit_log(original_ts, data):
    try:
        machine          = data.get("machine", "").strip()
        fault            = data.get("fault", "").strip()
        remedy           = data.get("remedy", "").strip()
        shift            = data.get("shift", "").strip()
        employee         = data.get("employee", "").strip()
        start_str        = data.get("startTime", "").strip()
        end_str          = data.get("endTime", "").strip()
        status           = data.get("status", "Running").strip()
        part_change      = data.get("partChange", "").strip()
        remarks          = data.get("remarks", "").strip()
        
        if not original_ts:
            return False, "Missing original log ID.", False, None, None
            
        if not all([machine, fault, remedy, shift, employee, start_str, end_str]):
            return False, "Please fill all required fields.", False, None, None
            
        fmt = "%Y-%m-%d %H:%M"
        try:
            start_dt = datetime.strptime(start_str, fmt)
            end_dt   = datetime.strptime(end_str,   fmt)
        except Exception:
            return False, "Invalid date/time format. Use YYYY-MM-DD HH:MM.", False, None, None
            
        if end_dt < start_dt:
            return False, "End time cannot be before start time.", False, None, None
            
        downtime = round((end_dt - start_dt).total_seconds() / 60, 2)
        
        wb = load_workbook(file_name)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        try:
            ts_idx = headers.index("Entry Timestamp")
        except ValueError:
            return False, "Entry Timestamp column not found.", False, None, None
            
        edit_cnt_idx = headers.index("Edit Count") if "Edit Count" in headers else None
        
        row_found = None
        for r_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r_idx, column=ts_idx + 1).value
            if cell_val and str(cell_val).strip() == original_ts:
                row_found = r_idx
                break
                
        if not row_found:
            return False, "Log entry not found.", False, None, None
            
        orig_edit_cnt = 0
        if edit_cnt_idx is not None:
            val = ws.cell(row=row_found, column=edit_cnt_idx + 1).value
            try:
                orig_edit_cnt = int(val or 0)
            except ValueError:
                orig_edit_cnt = 0
                
        # Perform updates
        def update_cell(col_name, val):
            if col_name in headers:
                col_i = headers.index(col_name) + 1
                ws.cell(row=row_found, column=col_i, value=val)
                
        update_cell("Machine Name", machine)
        update_cell("Fault", fault)
        update_cell("Remedy", remedy)
        update_cell("Shift", shift)
        update_cell("Employee Name", employee)
        update_cell("Fault Start Time", start_dt.strftime(fmt))
        update_cell("Fault End Time", end_dt.strftime(fmt))
        update_cell("Downtime", downtime)
        update_cell("Machine Status", status)
        update_cell("Part Change", part_change)
        update_cell("Remarks", remarks)
        update_cell("Edit Count", orig_edit_cnt + 1)
        
        wb.save(file_name)
        
        # Optional Postgres sync
        db_synced = False
        db_error = None
        if PSYCOPG2_AVAILABLE:
            try:
                conn = psycopg2.connect(
                    host="172.25.0.155", port=5432, database="Shed1",
                    user="postgres", password="12345678", connect_timeout=3
                )
                cur = conn.cursor()
                cur.execute("""
                    UPDATE shed1_log SET
                        machinename = %s, fault = %s, remedy = %s, shift = %s, employeename = %s,
                        faultstarttime = %s, faultendtime = %s, downtime = %s,
                        machinestatus = %s, partchange = %s, remarks = %s
                    WHERE entrytimestamp = %s
                """, (machine, fault, remedy, shift, employee,
                      start_dt.strftime(fmt), end_dt.strftime(fmt),
                      downtime, status, part_change, remarks, original_ts))
                conn.commit()
                conn.close()
                db_synced = True
            except Exception as e:
                db_error = str(e)
                
        entry_data = {
            "machine": machine, "fault": fault, "remedy": remedy,
            "shift": shift, "employee": employee, "downtime": downtime,
            "status": status, "partChange": part_change,
            "remarks": remarks, "timestamp": original_ts,
            "startTime": start_str, "endTime": end_str,
            "editCount": orig_edit_cnt + 1
        }
        return True, None, db_synced, db_error, entry_data
    except Exception as e:
        return False, str(e), False, None, None

# DELETE /api/logs — delete a log entry (admin only)
@app.route('/api/logs', methods=['DELETE'])
def delete_log():
    if not verify_admin(request):
        return jsonify({"error": "Unauthorized"}), 401
    
    timestamp = request.args.get("timestamp", "").strip()
    if not timestamp:
        data = request.get_json(silent=True) or {}
        timestamp = data.get("timestamp", "").strip()
        
    if not timestamp:
        return jsonify({"error": "Missing log timestamp"}), 400
        
    success, err_msg, db_synced, db_error = perform_delete_log(timestamp)
    if not success:
        return jsonify({"error": err_msg}), 404 if "not found" in err_msg.lower() else 500
        
    return jsonify({"success": True, "db_synced": db_synced, "db_error": db_error})

# POST /api/logs/edit — edit a log entry (admin only)
@app.route('/api/logs/edit', methods=['POST'])
def edit_log():
    if not verify_admin(request):
        return jsonify({"error": "Unauthorized"}), 401
        
    data             = request.json or {}
    original_ts      = data.get("originalTimestamp", "").strip()
    
    success, err_msg, db_synced, db_error, entry_data = perform_edit_log(original_ts, data)
    if not success:
        return jsonify({"error": err_msg}), 400 if "fill all" in err_msg or "format" in err_msg or "before" in err_msg else 404 if "not found" in err_msg else 500
        
    return jsonify({
        "success": True,
        "db_synced": db_synced,
        "db_error": db_error,
        "entry": entry_data
    })

# GET /api/requests — load pending requests (admin only)
@app.route('/api/requests', methods=['GET'])
def get_requests():
    if not verify_admin(request):
        return jsonify({"error": "Unauthorized"}), 401
    requests_list = load_requests()
    enriched_requests = []
    for req in requests_list:
        orig = get_log_by_timestamp(req.get("log_timestamp"))
        req["original_data"] = orig
        enriched_requests.append(req)
    return jsonify(enriched_requests)

# POST /api/requests — employee submits a request to edit/delete
@app.route('/api/requests', methods=['POST'])
def create_request():
    data = request.json or {}
    req_type = data.get("type", "").strip()
    log_ts = data.get("log_timestamp", "").strip()
    employee = data.get("employee", "").strip()
    
    if not req_type or not log_ts:
        return jsonify({"error": "Missing request type or log timestamp"}), 400
        
    if req_type not in ["edit", "delete"]:
        return jsonify({"error": "Invalid request type"}), 400
        
    # Check if target log exists
    orig = get_log_by_timestamp(log_ts)
    if not orig:
        return jsonify({"error": "Original log entry not found"}), 404
        
    req_id = f"req_{int(datetime.now().timestamp() * 1000)}"
    new_req = {
        "id": req_id,
        "type": req_type,
        "log_timestamp": log_ts,
        "employee": employee or orig.get("employee", ""),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "new_data": data.get("new_data") if req_type == "edit" else None
    }
    
    requests_list = load_requests()
    requests_list.append(new_req)
    save_requests(requests_list)
    
    return jsonify({"success": True, "request_id": req_id})

# POST /api/requests/action — admin approves or rejects request
@app.route('/api/requests/action', methods=['POST'])
def request_action():
    if not verify_admin(request):
        return jsonify({"error": "Unauthorized"}), 401
        
    data = request.json or {}
    req_id = data.get("request_id", "").strip()
    action = data.get("action", "").strip() # "approve" or "reject"
    
    if not req_id or action not in ["approve", "reject"]:
        return jsonify({"error": "Missing request_id or invalid action"}), 400
        
    requests_list = load_requests()
    target_req = None
    for req in requests_list:
        if req.get("id") == req_id:
            target_req = req
            break
            
    if not target_req:
        return jsonify({"error": "Request not found"}), 404
        
    if action == "reject":
        # Remove request
        requests_list = [r for r in requests_list if r.get("id") != req_id]
        save_requests(requests_list)
        return jsonify({"success": True})
        
    # If action is approve
    req_type = target_req.get("type")
    log_ts = target_req.get("log_timestamp")
    
    db_synced = False
    db_error = None
    
    if req_type == "delete":
        success, err_msg, db_synced, db_error = perform_delete_log(log_ts)
        if not success:
            return jsonify({"error": f"Failed to execute delete: {err_msg}"}), 500
            
    elif req_type == "edit":
        new_data = target_req.get("new_data", {})
        success, err_msg, db_synced, db_error, _ = perform_edit_log(log_ts, new_data)
        if not success:
            return jsonify({"error": f"Failed to execute edit: {err_msg}"}), 500
    else:
        return jsonify({"error": "Invalid request type"}), 400
        
    # If execution succeeded, remove request
    requests_list = [r for r in requests_list if r.get("id") != req_id]
    save_requests(requests_list)
    
    return jsonify({
        "success": True,
        "db_synced": db_synced,
        "db_error": db_error
    })

# GET /api/logs/today
@app.route('/api/logs/today', methods=['GET'])
def get_today_logs():
    if not os.path.exists(file_name):
        return jsonify([])
    try:
        wb = load_workbook(file_name, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return jsonify([])
        hi = {h: i for i, h in enumerate(headers) if h}

        def gv(row, col, idx):
            pos = hi.get(col, idx)
            return row[pos] if pos < len(row) else None

        today   = datetime.now().date()
        entries = []
        for row in rows:
            if not row or len(row) < 5:
                continue
            ts_str = gv(row, "Entry Timestamp", 8)
            dt = parse_dt(ts_str)
            if dt and dt.date() == today:
                entries.append({
                    "machine":    gv(row, "Machine Name",    0),
                    "fault":      gv(row, "Fault",           1),
                    "remedy":     gv(row, "Remedy",          2),
                    "shift":      gv(row, "Shift",           3),
                    "employee":   gv(row, "Employee Name",   4),
                    "startTime":  gv(row, "Fault Start Time",5),
                    "endTime":    gv(row, "Fault End Time",  6),
                    "downtime":   gv(row, "Downtime",        7),
                    "timestamp":  ts_str,
                    "status":     gv(row, "Machine Status",  9),
                    "partChange": gv(row, "Part Change",    10),
                    "remarks":    gv(row, "Remarks",        11),
                    "editCount":  int(gv(row, "Edit Count", 13) or 0),
                })
        entries.reverse()
        return jsonify(entries)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# GET /api/analytics
@app.route('/api/analytics', methods=['GET'])
def get_analytics():
    start_date_str = request.args.get("start_date")
    end_date_str   = request.args.get("end_date")
    if not start_date_str or not end_date_str:
        return jsonify({"error": "start_date and end_date required"}), 400
    try:
        start_filter = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_filter   = datetime.strptime(end_date_str,   "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        return jsonify({"error": "Use YYYY-MM-DD format"}), 400

    empty = {"summary": {"total_faults": 0, "total_downtime": 0, "top_fault": "None", "active_machines": 0},
             "trends": [], "faults": [], "machines": [], "shifts": [], "remedies": {}}

    if not os.path.exists(file_name):
        return jsonify(empty)

    try:
        wb = load_workbook(file_name, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return jsonify(empty)
        hi = {h: i for i, h in enumerate(headers) if h}

        def gv(row, col, idx):
            pos = hi.get(col, idx)
            return row[pos] if pos < len(row) else None

        logs = []
        for row in rows:
            if not row or not row[0]:
                continue
            st_str = gv(row, "Fault Start Time", 5)
            dt = parse_dt(st_str) or parse_dt(gv(row, "Entry Timestamp", 8))
            if dt and start_filter <= dt <= end_filter:
                logs.append({
                    "machine":  gv(row, "Machine Name", 0),
                    "fault":    gv(row, "Fault",        1),
                    "remedy":   gv(row, "Remedy",       2),
                    "shift":    gv(row, "Shift",        3),
                    "downtime": float(gv(row, "Downtime", 7) or 0),
                    "date":     dt,
                })

        total_faults   = len(logs)
        total_downtime = sum(l["downtime"] for l in logs)
        fault_counter  = Counter(l["fault"]   for l in logs)
        top_fault      = fault_counter.most_common(1)[0][0] if fault_counter else "None"
        unique_machines= {l["machine"] for l in logs}

        monthly_totals = defaultdict(int)
        monthly_faults = defaultdict(lambda: defaultdict(int))
        for l in logs:
            ym = l["date"].strftime("%Y-%m")
            monthly_totals[ym] += 1
            monthly_faults[ym][l["fault"]] += 1

        sorted_months   = sorted(monthly_totals)
        top_fault_names = [f[0] for f in fault_counter.most_common(5)]
        trends = []
        for ym in sorted_months:
            tot = monthly_totals[ym]
            shares = {fn: round(monthly_faults[ym].get(fn, 0) / tot * 100, 1) if tot else 0
                      for fn in top_fault_names}
            trends.append({"month": ym, "total_faults": tot, "shares": shares})

        faults_data   = sorted([{"name": n, "count": c} for n, c in fault_counter.items()],
                               key=lambda x: x["count"], reverse=True)
        machine_dt    = defaultdict(float)
        for l in logs:
            machine_dt[l["machine"]] += l["downtime"]
        machines_data = sorted([{"name": m, "downtime": round(d, 2)} for m, d in machine_dt.items()],
                               key=lambda x: x["downtime"], reverse=True)

        shift_counter = Counter(l["shift"] for l in logs)
        shifts_data   = [{"name": s, "count": c} for s, c in shift_counter.items()]

        remedy_map = defaultdict(Counter)
        for l in logs:
            if l["remedy"]:
                remedy_map[l["fault"]][l["remedy"]] += 1
        remedies_data = {fn: [{"name": n, "count": c} for n, c in rc.most_common(5)]
                         for fn, rc in remedy_map.items()}

        return jsonify({
            "summary":  {"total_faults": total_faults, "total_downtime": round(total_downtime, 2),
                         "top_fault": top_fault, "active_machines": len(unique_machines)},
            "trends":   trends,
            "faults":   faults_data,
            "machines": machines_data,
            "shifts":   shifts_data,
            "remedies": remedies_data,
        })
    except Exception as e:
        print("Analytics error:", e)
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
