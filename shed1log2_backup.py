import tkinter as tk
from tkinter import ttk, messagebox, font
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import os
import psycopg2
from PIL import Image, ImageTk
from datetime import date, timedelta
from calendar import monthrange
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from collections import Counter

# Set default fonts
DEFAULT_FONT = ('Arial', 10)
HEADING_FONT = ('timesnewroman', 12, 'bold')
TITLE_FONT = ('Arial', 30, 'bold')
BUTTON_FONT = ('Arial', 11, 'bold')

def enable_search_in_combobox(combo, values):
    def on_key_release(event):
        current_text = combo.get()
        if current_text:
            # Filter values that START with the typed text (case insensitive)
            filtered = [v for v in values if v.lower().startswith(current_text.lower())]
            if not filtered:
                # If no values start with text, show all that contain it
                filtered = [v for v in values if current_text.lower() in v.lower()]
        else:
            filtered = values
        
        combo['values'] = filtered
        
        # Auto-open dropdown to show options
        combo.event_generate('<Down>')

    def on_focus_out(event):
        combo['values'] = values

    combo.bind('<KeyRelease>', on_key_release)
    combo.bind('<FocusOut>', on_focus_out)

# File paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
machines_file = os.path.join(BASE_DIR, "machines.txt")
faults_file = os.path.join(BASE_DIR, "faults.txt")
remedies_file = os.path.join(BASE_DIR, "remedies.txt")
employees_file = os.path.join(BASE_DIR, "employees.txt")
parts_file = os.path.join(BASE_DIR, "parts.txt")
file_name = os.path.join(BASE_DIR, "machine_log.xlsx")


def ensure_file(file_path):
    if not os.path.exists(file_path):
        with open(file_path, "w") as f:
            f.write("")

def load_list(file_path):
    ensure_file(file_path)
    with open(file_path, "r") as f:
        return [line.strip() for line in f if line.strip()]

def initialize_excel_file():
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Machine Log"
        ws.append(["Machine Name", "Fault", "Remedy", "Shift", "Employee Name",
                   "Fault Start Time", "Fault End Time", "Downtime", "Entry Timestamp",
                   "Machine Status", "Part Change", "Remarks", "Fault Location"])
        wb.save(file_name)

initialize_excel_file()

# Load dropdown data
faults_list = load_list(faults_file)
remedies_list = load_list(remedies_file)
employees_list = load_list(employees_file)
machines_list = load_list(machines_file)
parts_list = load_list(parts_file)

# GUI Setup
root = tk.Tk()
root.title("Machine Fault Logging System")
root.state('zoomed')
root.configure(bg="#f0f0f0")

top_bar = tk.Frame(root, bg="#f0f0f0")
top_bar.pack(fill="x", padx=10, pady=10)

try:
    image = Image.open("logo.png")
    image = image.resize((350, 105))  # Adjust size if needed
    logo = ImageTk.PhotoImage(image)
    logo_label = tk.Label(top_bar, image=logo, bg="#f0f0f0")
    logo_label.image = logo
    logo_label.pack(side="left", padx=100, pady=0)
except:
    pass

# Title next to logo
tk.Label(top_bar, text="SHED-01 MACHINE LOG", font=TITLE_FONT, bg="#ffffff").pack(side="left", padx=200)

# Main container frame
form_frame = tk.Frame(root, bg="blue", bd=2, relief=tk.GROOVE)
form_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

# Canvas with vertical scrollbar
canvas = tk.Canvas(form_frame, bg="white")
scrollbar = ttk.Scrollbar(form_frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas, bg="white")

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)
canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# Split main form and today's entries
main_frame = tk.Frame(scrollable_frame, bg="white", bd=2, relief=tk.GROOVE)
main_frame.pack(side="left", fill="both", expand=True, padx=(0, 10))

right_frame = tk.Frame(scrollable_frame, bg="#f9f9f9", bd=2, relief=tk.GROOVE, width=800)
right_frame.pack(side="right", fill="y")
right_frame.pack_propagate(False)

# Today's Entries Title and List
tk.Label(right_frame, text="Today's Entries", font=HEADING_FONT, bg="#f9f9f9").pack(pady=10)
entry_listbox = tk.Listbox(right_frame, font=DEFAULT_FONT, width=45, height=30)
entry_listbox.pack(padx=10, pady=5, fill="both", expand=True)

def generate_graphs():
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        
        # Collect data
        faults_count = Counter()
        machines_downtime = {}
        employees_count = Counter()
        shifts_count = Counter()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Machine name
                machine = row[0]
                fault = row[1]
                downtime = row[7]
                employee = row[4]
                shift = row[3]
                
                faults_count[fault] += 1
                machines_downtime[machine] = machines_downtime.get(machine, 0) + (downtime or 0)
                employees_count[employee] += 1
                shifts_count[shift] += 1
        
        # Create figure with 4 subplots
        fig = Figure(figsize=(15, 10), dpi=100)
        
        # 1. Fault Frequency
        ax1 = fig.add_subplot(2, 2, 1)
        if faults_count:
            faults, counts = zip(*faults_count.most_common(10))
            ax1.barh(faults, counts, color='#FF6B6B')
            ax1.set_xlabel('Frequency')
            ax1.set_title('Top 10 Faults', fontweight='bold')
        
        # 2. Machine Downtime
        ax2 = fig.add_subplot(2, 2, 2)
        if machines_downtime:
            machines = list(machines_downtime.keys())[:10]
            downtime_vals = [machines_downtime[m] for m in machines]
            ax2.bar(machines, downtime_vals, color='#4ECDC4')
            ax2.set_ylabel('Downtime (minutes)')
            ax2.set_title('Machine Downtime', fontweight='bold')
            ax2.tick_params(axis='x', rotation=45)
        
        # 3. Employee Entries
        ax3 = fig.add_subplot(2, 2, 3)
        if employees_count:
            employees, counts = zip(*employees_count.most_common(10))
            ax3.bar(employees, counts, color='#95E1D3')
            ax3.set_ylabel('Number of Entries')
            ax3.set_title('Employee Entries', fontweight='bold')
            ax3.tick_params(axis='x', rotation=45)
        
        # 4. Shift Distribution
        ax4 = fig.add_subplot(2, 2, 4)
        if shifts_count:
            shifts_list = list(shifts_count.keys())
            shift_counts = list(shifts_count.values())
            colors = ['#FFD93D', '#6BCB77', '#4D96FF', '#FF6B9D']
            ax4.pie(shift_counts, labels=shifts_list, autopct='%1.1f%%', colors=colors[:len(shifts_list)])
            ax4.set_title('Shift Distribution', fontweight='bold')
        
        fig.tight_layout()
        return fig
    except Exception as e:
        messagebox.showerror("Graph Error", f"Failed to generate graphs: {e}")
        return None

def open_graphs_window():
    graphs_win = tk.Toplevel(root)
    graphs_win.title("Analytics Dashboard")
    graphs_win.geometry("1200x800")
    
    fig = generate_graphs()
    if fig:
        canvas = FigureCanvasTkAgg(fig, master=graphs_win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

def get_today_entries():
    today_entries = []
    try:
        wb = load_workbook(file_name)
        ws = wb.active
        today = datetime.now().date()
        for row in ws.iter_rows(min_row=2, values_only=True):
            entry_time = row[8]
            if entry_time:
                try:
                    entry_dt = datetime.strptime(entry_time, "%Y-%m-%d %H:%M:%S")
                    if entry_dt.date() == today:
                        today_entries.append(row)
                except:
                    continue
    except Exception as e:
        print("Failed to load entries:", e)
    return today_entries

def load_today_entries():
    entry_listbox.delete(0, tk.END)
    for entry in get_today_entries():
        entry_listbox.insert(tk.END, f"{entry[0]}  | {entry[9]} | {entry[10]} | {entry[11]}| {entry[12]} | {entry[1]} | {entry[4]} | {entry[2]} | {entry[3]}| {entry[5]} | {entry[6]}| {entry[8]} | {entry[7]}")

load_today_entries()

form_frame = tk.Frame(main_frame, bg="white")
form_frame.pack(fill="both", expand=True, padx=20, pady=20)

for i in range(3):
    form_frame.grid_columnconfigure(i, weight=1)

row = 0
def label_entry(label_text):
    global row
    label = tk.Label(form_frame, text=label_text, bg="white", anchor="w", font=DEFAULT_FONT)
    label.grid(row=row, column=0, sticky="e", padx=5, pady=10)
    entry = tk.Entry(form_frame, width=35, font=DEFAULT_FONT)
    entry.grid(row=row, column=1, padx=5, pady=10, sticky="ew")
    row += 1
    return entry

def combo_entry(label_text, values, readonly=False):
    global row
    tk.Label(form_frame, text=label_text, bg="white", anchor="w", font=DEFAULT_FONT).grid(
        row=row, column=0, sticky="e", padx=5, pady=10)
    state = "readonly" if readonly else "normal"
    combo = ttk.Combobox(form_frame, values=values, width=33, state=state, font=DEFAULT_FONT)
    combo.grid(row=row, column=1, padx=5, pady=10, sticky="ew")
    row += 1
    return combo

machine_cb = combo_entry("Machine Name", machines_list)
fault_cb = combo_entry("Fault", faults_list)
remedy_cb = combo_entry("Remedy", remedies_list)
shift_cb = combo_entry("Shift", ["General", "A", "B", "C"], readonly=True)
shift_cb.set("Select Shift")
status_cb = combo_entry("Machine Status", ["Running", "Stop"], readonly=True)
status_cb.set("Running")
employee_cb = combo_entry("Employee Name", employees_list)

enable_search_in_combobox(machine_cb, machines_list)
enable_search_in_combobox(fault_cb, faults_list)
enable_search_in_combobox(remedy_cb, remedies_list)
enable_search_in_combobox(employee_cb,employees_list)

employee_display = tk.Entry(form_frame, state="readonly", width=35, font=DEFAULT_FONT)
selected_employees = []
row += 1

part_change_cb = combo_entry("Part Change", parts_list)
enable_search_in_combobox(part_change_cb, parts_list)
remarks_entry = label_entry("Remarks")

current_year = 2025
mindate = date(current_year, 1, 1)
maxdate = date(current_year, 12, 31)

today = date.today()
first_day_current_month = today.replace(day=1)
last_day_prev_month = first_day_current_month - timedelta(days=1)
mindate = last_day_prev_month - timedelta(days=9)
last_day_current_month = today.replace(
    day=monthrange(today.year, today.month)[1]
)
maxdate = last_day_current_month

tk.Label(form_frame, text="Start Date", bg="white", font=DEFAULT_FONT).grid(
    row=row, column=0, padx=5, pady=10, sticky="e")
start_date = DateEntry(form_frame, date_pattern="mm/dd/yy", font=DEFAULT_FONT,
                       mindate=mindate, maxdate=maxdate, showothermonthdays=False)
start_date.grid(row=row, column=1, padx=5, pady=10, sticky="ew")
row += 1

tk.Label(form_frame, text="Start Time", bg="white", font=DEFAULT_FONT).grid(
    row=row, column=0, padx=5, pady=10, sticky="e")
time_frame = tk.Frame(form_frame, bg="white")
time_frame.grid(row=row, column=1, padx=5, pady=10, sticky="w")

start_hour_cb = ttk.Combobox(time_frame, values=["HH"] + [f"{i:02d}" for i in range(24)],
                             width=5, state='readonly', font=DEFAULT_FONT)
start_minute_cb = ttk.Combobox(time_frame, values=["MM"] + [f"{i:02d}" for i in range(60)],
                               width=5, state='readonly', font=DEFAULT_FONT)
start_hour_cb.pack(side="left", padx=(0, 5))
start_minute_cb.pack(side="left")
start_hour_cb.set("HH")
start_minute_cb.set("MM")
row += 1

tk.Label(form_frame, text="End Date", bg="white", font=DEFAULT_FONT).grid(
    row=row, column=0, padx=5, pady=10, sticky="e")
end_date = DateEntry(form_frame, date_pattern="mm/dd/yy", font=DEFAULT_FONT,
                     mindate=mindate, maxdate=maxdate)
end_date.grid(row=row, column=1, padx=5, pady=10, sticky="ew")
row += 1

tk.Label(form_frame, text="End Time", bg="white", font=DEFAULT_FONT).grid(
    row=row, column=0, padx=5, pady=10, sticky="e")
time_frame = tk.Frame(form_frame, bg="white")
time_frame.grid(row=row, column=1, padx=5, pady=10, sticky="w")

end_hour_cb = ttk.Combobox(time_frame, values=["HH"] + [f"{i:02d}" for i in range(24)],
                           width=5, state='readonly', font=DEFAULT_FONT)
end_minute_cb = ttk.Combobox(time_frame, values=["MM"] + [f"{i:02d}" for i in range(60)],
                             width=5, state='readonly', font=DEFAULT_FONT)
end_hour_cb.pack(side="left", padx=(0, 5))
end_minute_cb.pack(side="left")
end_hour_cb.set("HH")
end_minute_cb.set("MM")
row += 1

def clear_fields():
    machine_cb.set('')
    fault_cb.set('')
    remedy_cb.set('')
    shift_cb.set("Select Shift")
    status_cb.set("Running")
    part_change_cb.set('')
    remarks_entry.delete(0, tk.END)
    employee_display.config(state="normal")
    employee_display.delete(0, tk.END)
    employee_display.config(state="readonly")
    selected_employees.clear()
    start_date.set_date(datetime.now())
    end_date.set_date(datetime.now())
    start_hour_cb.set("HH")
    start_minute_cb.set("MM")
    end_hour_cb.set("HH")
    end_minute_cb.set("MM")

def save_data():
    try:
        machine = machine_cb.get().strip()
        fault = fault_cb.get().strip()
        remedy = remedy_cb.get().strip()
        shift = shift_cb.get().strip()
        status = status_cb.get().strip()
        part_change = part_change_cb.get().strip()
        remarks = remarks_entry.get().strip()
        employees = employee_cb.get().strip()

        if not (machine and fault and remedy and shift != "Select Shift" and employees):
            messagebox.showwarning("Missing Info", "Please fill all required fields.")
            return

        sh = start_hour_cb.get()
        sm = start_minute_cb.get()
        eh = end_hour_cb.get()
        em = end_minute_cb.get()

        if sh == "HH" or sm == "MM" or eh == "HH" or em == "MM":
            messagebox.showerror("Invalid Time", "Please select valid time.")
            return

        fmt = "%Y-%m-%d %H:%M"
        start_dt = datetime.strptime(f"{start_date.get_date()} {sh}:{sm}", fmt)
        end_dt = datetime.strptime(f"{end_date.get_date()} {eh}:{em}", fmt)
        if end_dt < start_dt:
            messagebox.showerror("Invalid Time", "End time cannot be before start time.")
            return

        downtime = round((end_dt - start_dt).total_seconds() / 60, 2)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Save to Excel
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([machine, fault, remedy, shift, employees,
                   start_dt.strftime(fmt), end_dt.strftime(fmt), downtime, timestamp,
                   status, part_change, remarks])
        wb.save(file_name)

        try:
            conn = psycopg2.connect(
                host="172.25.0.155",
                port=5432,
                database="Shed1",
                user="postgres",
                password="12345678"
            )
            cur = conn.cursor()
            cur.execute("""
            INSERT INTO shed1_log (
                machinename, fault, remedy, shift, employeename,
                faultstarttime, faultendtime, downtime, entrytimestamp,
                machinestatus, partchange, remarks
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                machine, fault, remedy, shift, employees,
                start_dt.strftime(fmt), end_dt.strftime(fmt), downtime, timestamp,
                status, part_change, remarks
            ))
            conn.commit()
            cur.close()
            conn.close()
            messagebox.showinfo("Saved", "Entry saved to Excel and Database successfully.")
        except Exception as db_error:
            messagebox.showwarning("Saved (Local Only)", f"Entry saved to Excel successfully.\n\nDatabase unavailable: {str(db_error)}")
        
        clear_fields()
        load_today_entries()

    except Exception as e:
        messagebox.showerror("Error", f"Failed to save:\n{e}")

# Bottom Buttons
button_frame = tk.Frame(main_frame, bg="white")
button_frame.pack(fill="x", pady=20, padx=20)

tk.Button(button_frame, text="Save Entry", command=save_data,
          bg="#4CAF50", fg="white", font=BUTTON_FONT, width=15, height=2).pack(side="left", padx=10)

tk.Button(button_frame, text="Clear Fields", command=clear_fields,
          bg="#607D8B", fg="white", font=BUTTON_FONT, width=15, height=2).pack(side="left", padx=10)

tk.Button(button_frame, text="Graphs", command=open_graphs_window,
          bg="#FF9800", fg="white", font=BUTTON_FONT, width=15, height=2).pack(side="left", padx=10)

def open_admin_panel():
    def check_password():
        if pw_entry.get() == "Am1122":
            pw_win.destroy()
            open_editor()
        else:
            messagebox.showerror("Denied", "Wrong password!")

    def open_editor():
        win = tk.Toplevel(root)
        win.title("Edit Lists")
        win.geometry("600x800")

        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        def create_tab(name, file):
            frame = tk.Frame(notebook, bg="white")
            notebook.add(frame, text=name)

            scrollbar = tk.Scrollbar(frame)
            scrollbar.pack(side="right", fill="y")

            lb = tk.Listbox(frame, height=15, width=50, font=DEFAULT_FONT,
                            yscrollcommand=scrollbar.set)
            lb.pack(fill="both", expand=True, padx=10, pady=10)

            items = load_list(file)
            for i in items:
                lb.insert(tk.END, i)

            entry_frame = tk.Frame(frame, bg="white")
            entry_frame.pack(fill="x", padx=10, pady=5)

            entry = tk.Entry(entry_frame, font=DEFAULT_FONT)
            entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

            def add():
                val = entry.get().strip()
                if val:
                    with open(file, "a") as f:
                        f.write(val + "\n")
                    lb.insert(tk.END, val)
                    entry.delete(0, tk.END)

            def remove():
                sel = lb.curselection()
                if sel:
                    val = lb.get(sel)
                    lb.delete(sel)
                    lines = [l.strip() for l in open(file) if l.strip() != val]
                    with open(file, "w") as f:
                        f.write("\n".join(lines))

            button_frame = tk.Frame(frame, bg="white")
            button_frame.pack(fill="x", padx=10, pady=10)

            tk.Button(button_frame, text="Add", command=add, font=BUTTON_FONT,
                      bg="#4CAF50", fg="white").pack(side="left", padx=5)
            tk.Button(button_frame, text="Remove", command=remove, font=BUTTON_FONT,
                      bg="#f44336", fg="white").pack(side="left", padx=5)

        create_tab("Machines", machines_file)
        create_tab("Faults", faults_file)
        create_tab("Remedies", remedies_file)
        create_tab("Employees", employees_file)
        create_tab("Parts", parts_file)

    pw_win = tk.Toplevel(root)
    pw_win.title("Admin Login")
    pw_win.geometry("300x200")

    frame = tk.Frame(pw_win, bg="white")
    frame.pack(fill="both", expand=True, padx=20, pady=20)

    tk.Label(frame, text="Enter Password", font=HEADING_FONT, bg="white").pack(pady=10)
    pw_entry = tk.Entry(frame, show="*", font=DEFAULT_FONT)
    pw_entry.pack(pady=10, fill="x")

    button_frame = tk.Frame(frame, bg="white")
    button_frame.pack(fill="x", pady=10)

    tk.Button(button_frame, text="Login", command=check_password, font=BUTTON_FONT,
              bg="#2196F3", fg="white").pack(side="right")

tk.Button(button_frame, text="Admin Panel", command=open_admin_panel,
          bg="#2196F3", fg="white", font=BUTTON_FONT, width=15, height=2).pack(side="right", padx=10)

def _on_mousewheel(event):
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
canvas.bind_all("<MouseWheel>", _on_mousewheel)

root.mainloop()
